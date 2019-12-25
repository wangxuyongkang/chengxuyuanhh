#https://www.qichacha.com/search?key=搜索col1
#headers = {
            # 'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3722.400 QQBrowser/10.5.3738.400'
        # }
from openpyxl import load_workbook
from urllib import parse
import requests,os,time
from lxml import etree
class qichacha(object):


    #从xlsx中取到要搜索的关键字 dict
    def operation_excel(self):
        dict = []
        book = load_workbook("中国保险行业信息披露系统.XLSX")
        sheet_names = book.get_sheet_names()
        print(sheet_names)
        sheet1 = book[sheet_names[0]]
        #xlsx 名
        print(sheet1.title)
        for row in sheet1.iter_rows(max_col=2, min_row=2):
            # print(list(row))
            info = [cell.value for cell in row]
            dict.append(info[0])
        # print(dict)
        return dict
    def start_request(self,url):
        '''开始起始任务'''
        print('正在爬取当前页面', url)
        response_text = self.send_request(url)
        print()
        if response_text:
            self.parse_data(response_text, url)
    #发送请求
    def send_request(self,url,headers=None,allow_redirects=True):
        """发送请求"""
        if headers == None:
            headers = {
                'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3722.400 QQBrowser/10.5.3738.400',
                'Cookie':'QCCSESSID=v4a18hvdurqjudtf2uf256am76; UM_distinctid=16c65dccfd63bf-030a725f45a249-34594974-1fa400-16c65dccfd856; Hm_lvt_3456bee468c83cc63fb5147f119f1075=1565077067; zg_did=%7B%22did%22%3A%20%2216c65dcd53d4a1-0a0aae0f43af74-34594974-1fa400-16c65dcd53efef%22%7D; hasShow=1; _uab_collina=156507706719754202739591; acw_tc=ddc2939c15650770416088074ed479f09a8e83d6b60566b7c0fb0e65cb; CNZZDATA1254842228=1809615403-1565076758-https%253A%252F%252Fwww.sogou.com%252F%7C1565082158; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201565082951750%2C%22updated%22%3A%201565083666203%2C%22info%22%3A%201565077067077%2C%22superProperty%22%3A%20%22%7B%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%7D%22%2C%22referrerDomain%22%3A%20%22%22%7D; Hm_lpvt_3456bee468c83cc63fb5147f119f1075=1565083669'
            }
        try:
            response = requests.get(url=url,headers=headers,verify=True)
            print('——————————————————',response.status_code)
            if response.status_code == 200:
                return response.text
            elif response.status_code == 302:
                if not allow_redirects:
                    return response.text,response.headers
        except Exception as err:
            print(err,'请求失败')
    #拿到详情地址
    def parse_data(self,text,currentUrl):
        etree_html = etree.HTML(text)
        items = etree_html.xpath('//tbody[@id="search-result"]')
        for item in items:
            #https://www.qichacha.com/firm_44fc7c8dd6531c1d4cacb5a173c25437.html
            url_detail = 'https://www.qichacha.com'+item.xpath('//tr[@class="frtrt "]/td[3]/a/@href')[0]
            print(url_detail)
            self.send_url_detail(url=url_detail)
    #发送详情地址的请求
    def send_url_detail(self,url,headers=None,allow_redirects=True):
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Safari/537.36 Core/1.70.3722.400 QQBrowser/10.5.3738.400',
            'Cookie': 'QCCSESSID=v4a18hvdurqjudtf2uf256am76; UM_distinctid=16c65dccfd63bf-030a725f45a249-34594974-1fa400-16c65dccfd856; Hm_lvt_3456bee468c83cc63fb5147f119f1075=1565077067; zg_did=%7B%22did%22%3A%20%2216c65dcd53d4a1-0a0aae0f43af74-34594974-1fa400-16c65dcd53efef%22%7D; hasShow=1; _uab_collina=156507706719754202739591; acw_tc=ddc2939c15650770416088074ed479f09a8e83d6b60566b7c0fb0e65cb; CNZZDATA1254842228=1809615403-1565076758-https%253A%252F%252Fwww.sogou.com%252F%7C1565082158; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201565082951750%2C%22updated%22%3A%201565083666203%2C%22info%22%3A%201565077067077%2C%22superProperty%22%3A%20%22%7B%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%7D%22%2C%22referrerDomain%22%3A%20%22%22%7D; Hm_lpvt_3456bee468c83cc63fb5147f119f1075=1565083669'
        }
        try:
            response = requests.get(url=url,headers=headers,verify=True)
            print('——————————————————',response.status_code)
            if response.status_code == 200:
                print('详情页',response.status_code)
                return self.html_detail(response.text)
        except Exception as err:
            print('请求错误',err)
    def html_detail(self,html):
        etree_html = etree.HTML(html)
        items = etree_html.xpath('//div[@id="company-top"]//div[@class="row"]')
        print(len(items))
        items_dict = {}
        for item in items:
            items_dict['img'] = item.xpath('//div[@class="logo"]/div[@class="imgkuang"]/img/@src')
            items_dict['companyVisit'] = item.xpath('//div[@class="logo"]//id[@class="companyVisit"]/text()')
            items_dict['content'] = item.xpath('//div[@class="content"]/div[@class="row title jk-tip"]/h1/text()')
            items_dict['rowtags'] = item.xpath('//div[@class="content"]//div[@class="row tags"]/span/text()')
            items_dict['phone'] = item.xpath('//div[@class="dcontent"]/div[@class="row"]/span[@class="fc "]//span[@class="cvlu"]/span/a/text()').replace('\n','').replace(' ','')
            items_dict['cdes'] = item.xpath('//div[@class="dcontent"]/div[@class="row"]/span[@class="cvlu "]/a/text()')

        print(items_dict)
if __name__ == '__main__':
    run = qichacha()
    data_url = run.operation_excel()
    data_url = ['https://www.qichacha.com/search?key=' + parse.quote(i) for i in data_url]
    for url in data_url:
        time.sleep(1)
        run.start_request(url=url)
