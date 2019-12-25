import requests,re,os,time
from urllib import parse
from lxml import etree
from openpyxl import Workbook
import json
from openpyxl import load_workbook

"""
运行环境 python3

需要安装的依赖包
pip3 install requests
pip3 install lxml
pip3 install openpyxl
https://www.jianshu.com/p/ce2ba7caa414

https://s.weibo.com/weibo?
q=%E6%88%90%E9%83%BD%E4%B8%83%E4%B8%AD%E7%BE%8E%E9%A3%9F
&wvr=6&b=1&Refer=SWeibo_box&page=1

https://s.weibo.com/weibo?
q=%E6%88%90%E9%83%BD%E4%B8%83%E4%B8%AD%E7%BE%8E%E9%A3%9F
&wvr=6&b=1&Refer=SWeibo_box&page=2
"""


class Weibo(object):

    def __init__(self):
        #表格名称
        self.excel_filename = 'weibo_data.xlsx'
        #获取excel文件，和操作表格
        self.excel_worker,self.sheet = self.create_excel()
        self.row = 1

    def create_excel(self):
        """创建excel，不存在则创建"""
        if not os.path.isfile(self.excel_filename):
            #创建excel操作对象
            excel_worker = Workbook()
            #返回excel中的sheet（默认返回第一个）
            sheet = excel_worker.active

            sheet.title = self.excel_filename.replace('.xlsx','')

            titles = ['username', 'content', 'publishtime', 'source', 'collectnum', 'transmitnum', 'commentnum', 'likenum']

            for index, title in enumerate(titles):
                print(index, title)
                #cell方法往excel中写入数据（前一个参数表示行，表一个参数表示列）
                sheet.cell(1, index + 1).value = title

            # save执行save方法保存写入的数据
            excel_worker.save(self.excel_filename)
            return excel_worker,sheet
            
        else:
            """excel存在"""
            excel_worker = load_workbook(self.excel_filename)
            # excel_worker.index()
            # excel_worker.get_sheet_by_name()
            sheet = excel_worker.active
            return excel_worker,sheet

    def start_request(self, url):
        """开始起始任务"""
        print('正在爬取当前页面',url)
        response_text = self.send_request(url)
        if response_text:
            self.parse_data(response_text,url)

    def send_request(self,url,params=None, headers=None,allow_redirects=True):
        """发送请求"""
        if headers == None:
            #如果没有传headers,则使用默认的headers
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36',
                'Cookie':'login_sid_t=ba43d429d4c68db236d4795e28c93488; cross_origin_proto=SSL; _s_tentry=passport.weibo.com; Apache=5237650840611.107.1564321280044; SINAGLOBAL=5237650840611.107.1564321280044; ULV=1564321280097:1:1:1:5237650840611.107.1564321280044:; UOR=,,www.baidu.com; WBtopGlobal_register_version=307744aa77dd5677; SSOLoginState=1564790800; un=18518753265; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WF7PqHBmhw4uiXSk_1RUrCU5JpX5K2hUgL.Fo-cehe7Soe41Kq2dJLoIf-LxK-L12BL1KqLxKML1h2LB-zLxKqL1-qLBoBLxKML1-2L1hBLxKqL1hnL1K2LxKqL1KnLB-qLxKMLBKnL12zLxK-LB.2L12qLxKML12eLB.9kShzN; ALF=1596501484; SCF=AkFf4MnIt0fIF_4ehaPj4oGUWsdA1vFaDsGCohPVB1CggSpjRxtx5jO7SmprfBQIT_d8fVuD8Mx_r8yCuTmuKAM.; SUB=_2A25wQwo9DeRhGeNI61ER9i3FwjqIHXVTOXz1rDV8PUNbmtAKLXHTkW9NSKLWD3q6LaY3VueORiQfQx1slR92SLVP; SUHB=0S7y2ael5zxe2j; un=18518753265; wvr=6; webim_unReadCount=%7B%22time%22%3A1564965499854%2C%22dm_pub_total%22%3A0%2C%22chat_group_client%22%3A0%2C%22allcountNum%22%3A17%2C%22msgbox%22%3A0%7D; WBStorage=edfd723f2928ec64|undefined'}
        try:
            response = requests.get(url=url, allow_redirects=allow_redirects,headers=headers, params=params,verify=True)
            if response.status_code == 200:
                return response.text
            elif response.status_code == 302:
                if not allow_redirects:
                    return response.text,response.headers

        except Exception as err:
            print(err,'请求失败')


    def parse_data(self,text,currentUrl):

        # with open('page.html','w') as file:
        #     file.write(text)

        etree_html = etree.HTML(text)

        items = etree_html.xpath('//div[@class="card-wrap"]/div[@class="card"]')
        print(len(items))

        for item in items:
            self.row+=1
            item_dict = {}
            #用户
            item_dict['username'] = self.extract_first(item.xpath('.//div[@class="content"]//a[@class="name"]/text()'))
            #将数据写入excel
            self.sheet.cell(self.row,1).value = item_dict['username']
            #内容
            item_dict['content'] = self.extract_first(item.xpath('.//div[@class="content"]/p[@class="txt"][1]//text()'),isjoin=True)
            self.sheet.cell(self.row, 2).value = item_dict['content']
            #时间
            item_dict['publishtime'] = self.extract_first(item.xpath('.//div[@class="content"]/p[@class="from"]/a[1]/text()'))
            self.sheet.cell(self.row, 3).value = item_dict['publishtime']
            #来源
            item_dict['source'] = self.extract_first(item.xpath('.//div[@class="content"]/p[@class="from"]/a[2]/text()'),default='未知')
            self.sheet.cell(self.row, 4).value = item_dict['source']
            #收藏
            item_dict['collectnum'] = self.extract_first(item.xpath('./div[@class="card-act"]/ul/li[1]/a/text()'),default=0)
            item_dict['collectnum'] = self.get_num(item_dict['collectnum'])
            self.sheet.cell(self.row, 5).value = item_dict['collectnum']
            #转发
            item_dict['transmitnum'] = self.extract_first(item.xpath('./div[@class="card-act"]/ul/li[2]/a/text()'),default=0)
            item_dict['transmitnum'] = self.get_num(item_dict['transmitnum'])
            self.sheet.cell(self.row, 6).value = item_dict['transmitnum']
            #评论
            item_dict['commentnum'] = self.extract_first(item.xpath('./div[@class="card-act"]/ul/li[3]/a/text()'),default=0)
            item_dict['commentnum'] = self.get_num(item_dict['commentnum'])
            self.sheet.cell(self.row, 7).value = item_dict['commentnum']
            #点赞
            item_dict['likenum'] = self.extract_first(item.xpath('./div[@class="card-act"]/ul/li[4]/a/em/text()'),default=0)
            item_dict['likenum'] = int(item_dict['likenum'])
            self.sheet.cell(self.row, 8).value = item_dict['likenum']
            print(item_dict)

            #获取个人主页

            detail_url = 'https:'+self.extract_first(item.xpath('.//div[@class="content"]//a[@class="name"]/@href'))
            print(detail_url)
            text,headers = self.send_request(detail_url,allow_redirects=False)
            detail_url = parse.urljoin('https://weibo.com/ycrsksw?refer_flag=1001030103_&is_hot=1',headers['Location'])
            print(detail_url)

            profile_html = self.send_request(detail_url)
            # with open('profile_html.html','w') as file:
            #     file.write(profile_html)
            # profile_etree = etree.HTML(profile_html)
            # table = profile_etree.xpath('//table[@class="tb_counter"]')
            # print(len(table))
            pattern = re.compile('<script>FM.view\((.*?)\)</script>',re.S)
            result = re.findall(pattern,profile_html)
            if len(result) > 0:
                for item in result:
                    item = item.replace('<script>FM.view(','').replace(')</script>','').replace('<!---欧盟协议---->')
                    json_data = json.loads(item)
                    html = json_data.get('html','')


        # self.excel_worker.save(self.excel_filename)

        #获取下一页：
        next = etree_html.xpath('//div[@class="m-page"]//a[@class="next"][1]/@href')
        if next:
            next = parse.urljoin(currentUrl,next[0])
            time.sleep(5)
            self.start_request(next)


    def extract_first(self,data,default=None,isjoin=False):
        #获取结果处理
        if len(data) == 0:
            return default
        else:
            if isjoin:
                return ''.join(data).replace(' ','').replace('\n','')
            else:
                return data[0].replace(' ','').replace('\n','')

    def get_num(self,dataStr):
        """获取收藏、转发、评论的数字"""
        pattern = re.compile('\d+',re.S)
        result = re.search(pattern,dataStr)

        if result:
            return int(result.group())
        else:
            return 0


if __name__ == '__main__':

    """
    q: 成都七中食品
    wvr: 6
    b: 1
    Refer: SWeibo_box
    page: 1
    
    https://s.weibo.com/weibo?q=%E6%88%90%E9%83%BD%E4%B8%83%E4%B8%AD%E9%A3%9F%E5%93%81
    &wvr=6&b=1&Refer=SWeibo_box&page=1
    """
    #设置搜索关键字
    key = parse.quote('成都七中食品')
    # parse.urlencode()
    #根据关键字拼接起始url地址
    start_url = 'https://s.weibo.com/weibo?q=%s&wvr=6&b=1&Refer=SWeibo_box&page=1' % (key,)

    wb_spider = Weibo()
    wb_spider.start_request(url=start_url)








