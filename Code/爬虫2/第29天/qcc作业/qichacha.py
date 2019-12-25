# -*- coding: utf-8 -*-
import requests,os
from urllib import parse
from lxml.html import etree
from openpyxl import load_workbook

class QichachaSpider(object):

    def __init__(self,filename):
        self.name = 'qichacha'
        self.excel_filename = filename

    """
    step:根据https://www.qichacha.com/找到分类的信息,存入数据库
    """
    def start_requests(self):
        # keywords = self.read_excel()
        #https://www.qichacha.com/search?key=%E5%B0%8F%E7%B1%B3
        url = 'https://www.qichacha.com/search?key='
        # parse.urlencode()
        # parse.quote()
        keywords = self.read_excel()

        if keywords:
            for word in keywords:
                word = word.replace('.','')
                print(word)
                word = parse.quote(word) # 小米=>%E5%B0%8F%E7%B1%B3
                full_url = url + word
                print(full_url)

                html = self.send_request(full_url)
                if html:
                    self.parse_company_list(html)
        else:
            print('没有请求任务')

    def read_excel(self):
        if not os.path.isfile(self.excel_filename.replace('.xlsx','')):
            excel = load_workbook(self.excel_filename)
            # 获取sheet：
            table = excel.get_sheet_by_name('Data')  # 通过表名获取
            # 获取行数和列数：
            rows = table.max_row  # 获取行数
            cols = table.max_column  # 获取列数
            # 获取单元格值：
            datas = []
            for row in range(2,rows+1):
                value = table.cell(row=row, column=1).value
                datas.append(value)
            # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
            return datas


    # def get_proxy(self):
    #     import random
    #     proxies = [
    #         '192.168.1.11:8000','192.168.1.11','192.168.1.11',
    #         '192.168.1.11','192.168.1.11'
    #     ]
    #
    #     return {'https':random.choice(proxies)}

    def send_request(self, url, params=None, headers=None, allow_redirects=True,Refere=None):
        """发送请求"""
        if headers == None:
            # 如果没有传headers,则使用默认的headers
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
                'Cookie': 'QCCSESSID=gu9t272no1ll96orq9bi1ol224; UM_distinctid=16c65d5e993458-09f68f96b26654-37647e05-13c680-16c65d5e994369; zg_did=%7B%22did%22%3A%20%2216c65d5fb8e35-0bc499b4c21d3f-37647e05-13c680-16c65d5fb8f6d2%22%7D; hasShow=1; _uab_collina=156507662172351051399105; acw_tc=ddc2939615650766219376005e5f14197818c62443b8718681f67bec86; CNZZDATA1254842228=1806396661-1565071358-https%253A%252F%252Fsp0.baidu.com%252F%7C1565101458; Hm_lvt_3456bee468c83cc63fb5147f119f1075=1565076622,1565104357; zg_de1d1a35bfa24ce29bbf2c7eb17e6c4f=%7B%22sid%22%3A%201565104356917%2C%22updated%22%3A%201565104468019%2C%22info%22%3A%201565076618140%2C%22superProperty%22%3A%20%22%7B%7D%22%2C%22platform%22%3A%20%22%7B%7D%22%2C%22utm%22%3A%20%22%7B%7D%22%2C%22referrerDomain%22%3A%20%22sp0.baidu.com%22%2C%22cuid%22%3A%20%22c1f38c4992c3d367353906ba0b6242af%22%7D; Hm_lpvt_3456bee468c83cc63fb5147f119f1075=1565104469',
            }
            if Refere:
                headers['referer'] = Refere
        try:
            response = requests.get(url=url, allow_redirects=allow_redirects, headers=headers, params=params,
                                    verify=True,proxies=self.get_proxy())
            if response.status_code == 200:
                return response.text
            elif response.status_code == 302:
                if not allow_redirects:
                    return response.text, response.headers
        except Exception as err:
            print(err, '请求失败')


    def parse_company_list(self, html):
        """
        获取搜索结果中的第一条公司信息，并发起请求，获取页面源码
        :param html: 请求的页面源码
        :return:
        """

        response = etree.HTML(html)

        # company_trs = response.xpath('//tbody[@id="search-result"]//tr[@class="frtrt"]')
        # print(len(company_trs))
        # if len(company_trs) == 0:
        company_trs = response.xpath('//tbody[@id="search-result"]//tr[1]')

        print('获取公司搜索列表',len(company_trs))
        if len(company_trs) > 0:
            company = company_trs[0]
            item = {}
            # print(company)
            item['email'] = self.extract_first(company.xpath('./td[3]/p[@class="m-t-xs"][2]/text()'),'').replace(' ','').replace('邮箱：','').replace('\n','')
            item['phonenum'] = self.extract_first(company.xpath('./td[3]/p[@class="m-t-xs"][2]/span[@class="m-l"]/text()'),'').replace(' ','').replace('电话：','')
            item['companyName'] = self.extract_first(company.xpath('./td[3]/a[1]/em/em/text()'),'')
            item['lagal'] = self.extract_first(company.xpath('./td[3]/p[@class="m-t-xs"][1]/a/text()'),'')
            item['capital'] = self.extract_first(company.xpath('./td[3]/p[@class="m-t-xs"][1]/span[1]/text()'),'').replace('注册资本：','')
            item['buildDate'] = self.extract_first(company.xpath('./td[3]/p[@class="m-t-xs"][1]/span[2]/text()'),'').replace('成立日期：','')
            item['icon'] = self.extract_first(company.xpath('./td[@class="imgtd"]/img/@src'),'')
            item['address'] = self.extract_first(company.xpath('//p[@class="m-t-xs"][3]/text()'),'').replace(' ','').replace('\n','')
            item['status'] = self.extract_first(company.xpath('.//span[@class="nstatus text-success-lt m-l-xs"]/text()'),'')
            detail_url = self.extract_first(company.xpath('./td[3]/a[1]/@href'),'')

            detail_url = parse.urljoin('https://www.qichacha.com/firm_0f8cd2784b03d78c28d7ab916259d7d3.html',detail_url)
            print(detail_url)

            html = self.send_request(detail_url)
            if html:
                self.parse_company_detail(html,item)

    def parse_company_detail(self, html,item):
        """
        解析公司详情的数据
        :param response: 公司详情的响应结果
        :return:
        """
        # print('正在解析公司详情')

        # with open('detail.html','w') as file:
        #     file.write(html)

        response = etree.HTML(html)
        # 官网
        item['website'] = self.extract_first(response.xpath('//div[@class="dcontent"]//span[@class="cvlu "]/a/text()'),"")
        # 简介
        item['desc'] = self.extract_first(response.xpath('//div[@class="m-t-sm m-b-sm"]/pre/text()'),'')

        trs = response.xpath('//section[@id="Cominfo"]/table[@class="ntable"][1]//tr')
        if len(trs):
            # 经营状态
            item['scopeStatus'] = self.deal_with_str(trs[1].xpath('./td[2]/text()'))
            # 统一社会信用代码
            item['creditCode'] = self.deal_with_str(trs[2].xpath('./td[2]/text()'))
            # 注册号
            item['registNumber'] = self.deal_with_str(trs[3].xpath('./td[2]/text()'))
            # 企业类型
            item['companyType'] = self.deal_with_str(trs[4].xpath('./td[2]/text()'))
            # 核准日期
            item['checkTime'] = self.deal_with_str(trs[5].xpath('./td[2]/text()'))
            # 所属地区
            item['place_origin'] = self.deal_with_str(trs[6].xpath('./td[2]/text()'))
            # 曾用名
            item['oldName'] = self.deal_with_str(trs[7].xpath('./td[2]//text()'))
            # 人员规模
            item['person_number'] = self.deal_with_str(trs[9].xpath('./td[2]/text()'))
            # 实缴资本
            item['relcapital'] = self.deal_with_str(trs[0].xpath('./td[4]/text()'))
            # 纳税人识别号
            item['ratepayerCode'] = self.deal_with_str(trs[2].xpath('./td[4]/text()'))
            # 组织机构代码
            item['institutionalNumber'] = self.deal_with_str(trs[3].xpath('./td[4]/text()'))
            # 所属行业
            item['industry'] = self.deal_with_str(trs[4].xpath('./td[4]/text()'))
            # 登记机关
            item['registration_authority'] = self.deal_with_str(trs[5].xpath('./td[4]/text()'))
            # 英文名
            item['englishName'] = self.deal_with_str(trs[6].xpath('./td[4]/text()'))
            # 参保人数
            item['insuredNumber'] = self.deal_with_str(trs[7].xpath('./td[4]/text()'))
            # 营业期限
            item['business_term'] = self.deal_with_str(trs[8].xpath('./td[4]/text()'))
            # 企业地址
            item['company_adress'] = self.deal_with_str(trs[9].xpath('./td[2]/text()'))
            # 经营范围
            item['scope'] = self.deal_with_str(trs[10].xpath('./td[2]/text()'))

        print(item)

    
    def extract_first(self,data,default=None):
        if len(data) > 0:
            return data[0]
        return default
        
    
    def deal_with_str(self, dataarr):

        str = ''.join(dataarr).replace('\n', '').replace(' ', '')

        return str

if __name__ == '__main__':

    filename = '中国保险行业信息披露系统.xlsx'
    spider = QichachaSpider(filename)
    spider.start_requests()